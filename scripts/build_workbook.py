#!/usr/bin/env python3
"""Build the XtraLight CEU L&L Command Workbook from data/targets.csv.

Usage:  python scripts/build_workbook.py
Output: output/XtraLight_CEU_LL_Command_Workbook.xlsx

The CSV is the single source of truth. Edit it (in GitHub's web editor or
locally), rerun this script (or let the GitHub Action run it), and the
workbook - Dashboard included - is regenerated.
"""
import csv
import os
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.chart import BarChart, Reference

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(ROOT, "data", "targets.csv")
OUT_PATH = os.path.join(ROOT, "output", "XtraLight_CEU_LL_Command_Workbook.xlsx")

XLB = "0079C1"; GOLD = "C49F06"; GREY = "595959"
LIGHT = "F2F7FB"; INPUT = "FFF6DE"
GREEN1 = "C6EFCE"; GREEN2 = "A9D08E"; GREYF = "D9D9D9"
thin = Side(style="thin", color="D9D9D9")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUSES = ["Not Started", "Outreach Sent", "In Scheduling", "Scheduled",
            "Presented", "Follow-Up", "Declined", "Dormant"]

N_COLS = 21  # A..U


def F(sz=10, b=False, col="000000", it=False):
    return Font(name="Arial", size=sz, bold=b, color=col, italic=it)


def load_rows():
    with open(CSV_PATH, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = [(r + [""] * N_COLS)[:N_COLS] for r in reader if len(r) > 1 and r[1].strip()]
    rows.sort(key=lambda r: (r[0], r[1]))
    return header, rows


def build():
    header, rows = load_rows()
    wb = Workbook()

    # ---------------- Targets ----------------
    tg = wb.create_sheet("Targets")
    widths = [16, 30, 22, 22, 20, 30, 36, 10, 9, 54, 20, 14, 13, 18, 18, 18, 22, 12, 12, 24, 28]
    for i, (h, w) in enumerate(zip(header, widths), 1):
        c = tg.cell(row=1, column=i, value=h)
        c.font = F(10, True, "FFFFFF"); c.fill = PatternFill("solid", fgColor=XLB)
        c.alignment = Alignment(wrap_text=True, vertical="center"); c.border = BORD
        tg.column_dimensions[get_column_letter(i)].width = w
    tg.row_dimensions[1].height = 30

    input_fill = PatternFill("solid", fgColor=INPUT)
    t1_fill = PatternFill("solid", fgColor=LIGHT)
    for r_idx, row in enumerate(rows, 2):
        if not row[11].strip():
            row[11] = "Not Started"
        for c_idx, val in enumerate(row, 1):
            c = tg.cell(row=r_idx, column=c_idx, value=val)
            c.font = F(9, b=(c_idx == 2 and row[8] == "Tier 1"))
            c.alignment = Alignment(wrap_text=True, vertical="top"); c.border = BORD
            if c_idx <= 10 and row[8] == "Tier 1":
                c.fill = t1_fill
            if c_idx >= 11:
                c.fill = input_fill
    tg.auto_filter.ref = f"A1:U{len(rows) + 1}"
    tg.freeze_panes = "C2"

    dv_status = DataValidation(type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=True)
    tg.add_data_validation(dv_status); dv_status.add("L2:L1000")
    dv_course = DataValidation(type="list", formula1='"FED,COR,Either"', allow_blank=True)
    tg.add_data_validation(dv_course); dv_course.add("M2:M1000")

    tg.conditional_formatting.add("L2:L1000",
        CellIsRule(operator="equal", formula=['"Scheduled"'], fill=PatternFill("solid", fgColor=GREEN1)))
    tg.conditional_formatting.add("L2:L1000",
        CellIsRule(operator="equal", formula=['"Presented"'], fill=PatternFill("solid", fgColor=GREEN2)))
    tg.conditional_formatting.add("L2:L1000",
        CellIsRule(operator="equal", formula=['"Declined"'], fill=PatternFill("solid", fgColor=GREYF)))

    # ---------------- Dashboard ----------------
    db = wb.active; db.title = "Dashboard"
    db.sheet_view.showGridLines = False
    db["B2"] = "CEU LUNCH & LEARN COMMAND DASHBOARD"; db["B2"].font = F(16, True, XLB)
    db["B3"] = "XtraLight AIA CES - Government Sector Targets  |  All numbers update live from the Targets tab"
    db["B3"].font = F(9, False, GREY, it=True)

    kpis = [
        ("Total Targets",   "=COUNTA(Targets!$B$2:$B$1000)", "0"),
        ("Tier 1",          '=COUNTIF(Targets!$I$2:$I$1000,"Tier 1")', "0"),
        ("Not Started",     '=COUNTIF(Targets!$L$2:$L$1000,"Not Started")', "0"),
        ("Active Outreach", '=COUNTIF(Targets!$L$2:$L$1000,"Outreach Sent")+COUNTIF(Targets!$L$2:$L$1000,"In Scheduling")', "0"),
        ("Scheduled",       '=COUNTIF(Targets!$L$2:$L$1000,"Scheduled")', "0"),
        ("Presented",       '=COUNTIF(Targets!$L$2:$L$1000,"Presented")', "0"),
        ("Booked Rate",     "=IFERROR((F6+G6)/MAX(1,B6-D6),0)", "0%"),
    ]
    for i, (lab, form, fmt) in enumerate(kpis):
        col = get_column_letter(2 + i)
        lc = db[f"{col}5"]; lc.value = lab; lc.font = F(9, True, GREY)
        vc = db[f"{col}6"]; vc.value = form; vc.font = F(18, True, XLB); vc.number_format = fmt
        db.column_dimensions[col].width = 15

    db["B8"] = "Booking goal (edit):"; db["B8"].font = F(10, True)
    g = db["C8"]; g.value = 12; g.font = F(12, True, "0000FF"); g.fill = PatternFill("solid", fgColor=INPUT)
    db["D8"] = "Progress to goal:"; db["D8"].font = F(10, True)
    p = db["E8"]; p.value = "=MIN(1,(F6+G6)/MAX(1,C8))"; p.font = F(12, True, GOLD); p.number_format = "0%"
    db["F8"] = "(goal = Scheduled + Presented)"; db["F8"].font = F(8, False, GREY, it=True)

    db["B11"] = "PIPELINE BY STATUS"; db["B11"].font = F(11, True, XLB)
    for ref, t in (("B12", "Status"), ("C12", "Count"), ("D12", "% of Total")):
        db[ref] = t; db[ref].font = F(9, True, "FFFFFF")
        db[ref].fill = PatternFill("solid", fgColor=XLB); db[ref].border = BORD
    for i, s in enumerate(STATUSES):
        r = 13 + i
        db[f"B{r}"] = s; db[f"B{r}"].font = F(9); db[f"B{r}"].border = BORD
        db[f"C{r}"] = f'=COUNTIF(Targets!$L$2:$L$1000,"{s}")'
        db[f"C{r}"].font = F(9, True); db[f"C{r}"].border = BORD
        db[f"D{r}"] = f"=IFERROR(C{r}/MAX(1,$B$6),0)"
        db[f"D{r}"].font = F(9); db[f"D{r}"].number_format = "0%"; db[f"D{r}"].border = BORD
    db.conditional_formatting.add(f"C13:C{12 + len(STATUSES)}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=60, color=XLB, showValue=True))

    db["B23"] = "BY PRIORITY TIER"; db["B23"].font = F(11, True, XLB)
    for ref, t in (("B24", "Tier"), ("C24", "Firms"), ("D24", "Booked")):
        db[ref] = t; db[ref].font = F(9, True, "FFFFFF")
        db[ref].fill = PatternFill("solid", fgColor=XLB); db[ref].border = BORD
    for i, t in enumerate(["Tier 1", "Tier 2", "Tier 3"]):
        r = 25 + i
        db[f"B{r}"] = t; db[f"B{r}"].font = F(9); db[f"B{r}"].border = BORD
        db[f"C{r}"] = f'=COUNTIF(Targets!$I$2:$I$1000,"{t}")'
        db[f"C{r}"].font = F(9, True); db[f"C{r}"].border = BORD
        db[f"D{r}"] = (f'=COUNTIFS(Targets!$I$2:$I$1000,"{t}",Targets!$L$2:$L$1000,"Scheduled")'
                       f'+COUNTIFS(Targets!$I$2:$I$1000,"{t}",Targets!$L$2:$L$1000,"Presented")')
        db[f"D{r}"].font = F(9); db[f"D{r}"].border = BORD

    db["B30"] = "BY CEU COURSE FIT"; db["B30"].font = F(11, True, XLB)
    for ref, t in (("B31", "Fit"), ("C31", "Firms")):
        db[ref] = t; db[ref].font = F(9, True, "FFFFFF")
        db[ref].fill = PatternFill("solid", fgColor=XLB); db[ref].border = BORD
    for i, (lab, code) in enumerate([("FED (Federal Secure)", "FED"),
                                     ("COR (Correctional)", "COR"),
                                     ("Both courses", "FED+COR")]):
        r = 32 + i
        db[f"B{r}"] = lab; db[f"B{r}"].font = F(9); db[f"B{r}"].border = BORD
        db[f"C{r}"] = f'=COUNTIF(Targets!$H$2:$H$1000,"{code}")'
        db[f"C{r}"].font = F(9, True); db[f"C{r}"].border = BORD

    state_counts = Counter(r[0] for r in rows)
    states_sorted = sorted(state_counts, key=lambda s: (-state_counts[s], s))
    db["N2"] = "BY STATE (HQ)"; db["N2"].font = F(11, True, XLB)
    for j, h in enumerate(["State", "Firms", "Tier 1", "Booked"]):
        c = db.cell(row=3, column=14 + j, value=h)
        c.font = F(9, True, "FFFFFF"); c.fill = PatternFill("solid", fgColor=XLB); c.border = BORD
    for i, st in enumerate(states_sorted):
        r = 4 + i
        db.cell(row=r, column=14, value=st).font = F(9)
        db.cell(row=r, column=15, value=f"=COUNTIF(Targets!$A$2:$A$1000,$N{r})").font = F(9, True)
        db.cell(row=r, column=16,
                value=f'=COUNTIFS(Targets!$A$2:$A$1000,$N{r},Targets!$I$2:$I$1000,"Tier 1")').font = F(9)
        db.cell(row=r, column=17,
                value=(f'=COUNTIFS(Targets!$A$2:$A$1000,$N{r},Targets!$L$2:$L$1000,"Scheduled")'
                       f'+COUNTIFS(Targets!$A$2:$A$1000,$N{r},Targets!$L$2:$L$1000,"Presented")')).font = F(9)
        for cc in range(14, 18):
            db.cell(row=r, column=cc).border = BORD
    db.conditional_formatting.add(f"O4:O{3 + len(states_sorted)}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=12, color=GOLD, showValue=True))
    db.column_dimensions["N"].width = 20
    for col in ("O", "P", "Q"):
        db.column_dimensions[col].width = 8

    chart = BarChart(); chart.type = "bar"; chart.style = 10
    chart.title = "Pipeline by Status"
    data = Reference(db, min_col=3, min_row=12, max_row=12 + len(STATUSES))
    cats = Reference(db, min_col=2, min_row=13, max_row=12 + len(STATUSES))
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    chart.legend = None; chart.width = 13; chart.height = 8.5
    db.add_chart(chart, "G11")
    db.column_dimensions["A"].width = 2

    # ---------------- READ ME ----------------
    rm = wb.create_sheet("READ ME")
    rm.sheet_view.showGridLines = False

    def put(ref, text, b=False, sz=10, col="000000", it=False):
        c = rm[ref]; c.value = text; c.font = F(sz, b, col, it)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    put("B2", "XtraLight AIA CES Lunch & Learn - Command Workbook", True, 14, XLB)
    put("B3", "GENERATED FILE - built from data/targets.csv by scripts/build_workbook.py. "
              "To change the roster permanently, edit the CSV and rebuild (or commit to GitHub and let the "
              "Action rebuild). Direct edits to this file can be pulled back into the CSV with "
              "scripts/sync_from_workbook.py.", False, 9, "000000", True)
    put("B5", "WHAT TO EDIT", True, 11, XLB)
    put("B6", "Pale-yellow cells: columns K-U on the Targets tab (rep agency, status, course, office, contact, "
              "dates, actions, notes) and the blue Booking Goal cell on the Dashboard. Everything else is data "
              "or formulas - leave it alone and the Dashboard stays live.")
    put("B7", "The AE Works row is pre-filled as a format example - overwrite it with real data. Status and "
              "Course columns have dropdowns; pick from the list so the Dashboard counts stay accurate.")
    put("B8", "Status ladder: Not Started > Outreach Sent > In Scheduling > Scheduled > Presented, with "
              "Follow-Up / Declined / Dormant as exits. Scheduled turns green, Presented dark green, "
              "Declined gray - automatically.")
    put("B10", "HOW TO WORK THE LIST", True, 11, XLB)
    put("B11", "1. HQ state understates reach - most Tier 1-3 firms run 10-100+ offices. Book the office "
               "nearest the rep agency, not the HQ, except the Pittsburgh home-field targets.")
    put("B12", "2. Entry paths by hit rate: (a) rep agency's existing relationship with the office's electrical "
               "dept head or spec team; (b) the firm's published lunch-and-learn / CES presentation request "
               "route; (c) named electrical engineers / lighting designers via LinkedIn, citing the specific "
               "HSW course title and 1.0 LU|HSW credit.")
    put("B13", "3. Named individual contacts are deliberately not pre-loaded - they churn fast. Capture them in "
               "columns O-Q as you verify at booking time.")
    put("B14", "4. Filter plays: Priority = Tier 1 + CEU Fit = COR gives the corrections call block. "
               "State = PA gives the home-field week. State = NE gives the Omaha VA cluster - one trip, "
               "eight firms.")
    put("B16", "LEGEND", True, 11, XLB)
    put("B17", "Tier 1 - Book directly. Tier 2 - Second wave. Tier 3 - Giants / low-spec-leverage: "
               "office-by-office through reps only.")
    put("B18", "CEU Fit: FED = Federal Secure Facilities HSW. COR = Correctional/Secure Environments HSW. "
               "FED+COR = both. (Food & Bev course targets a different firm universe; SSOE is the one "
               "crossover flag.)")
    rm.column_dimensions["A"].width = 2
    rm.column_dimensions["B"].width = 120
    for r in (3, 6, 12, 14):
        rm.row_dimensions[r].height = 40

    # Force Excel to recalculate all formulas on open (openpyxl writes no cached values)
    wb.calculation.fullCalcOnLoad = True

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Built {OUT_PATH}: {len(rows)} firms, {len(states_sorted)} states")


if __name__ == "__main__":
    build()

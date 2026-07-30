#!/usr/bin/env python3
"""Pull edits made directly in the workbook back into data/targets.csv.

Usage:  python scripts/sync_from_workbook.py [path/to/workbook.xlsx]
        (defaults to output/XtraLight_CEU_LL_Command_Workbook.xlsx)

Use this when you've been working the Targets tab in Excel (statuses,
contacts, dates, notes) and want the CSV - the source of truth - to catch up
before committing. Then commit BOTH files.
"""
import csv
import os
import sys
import warnings

warnings.filterwarnings("ignore")
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(ROOT, "data", "targets.csv")
DEFAULT_XLSX = os.path.join(ROOT, "output", "XtraLight_CEU_LL_Command_Workbook.xlsx")

N_COLS = 21


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx):
        sys.exit(f"Workbook not found: {xlsx}")

    wb = load_workbook(xlsx)
    if "Targets" not in wb.sheetnames:
        sys.exit("No 'Targets' sheet found - is this the right file?")
    ws = wb["Targets"]

    rows = []
    for r in ws.iter_rows(min_row=1, max_col=N_COLS, values_only=True):
        if r[1] is None or str(r[1]).strip() == "":
            continue  # skip blank rows; header passes because col B = "Firm"
        rows.append(["" if v is None else str(v) for v in r])

    if len(rows) < 2:
        sys.exit("Refusing to write: fewer than 2 rows read from the workbook.")

    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)
    print(f"Synced {len(rows) - 1} firms from {os.path.basename(xlsx)} -> data/targets.csv")
    print("Now commit both data/targets.csv and the workbook.")


if __name__ == "__main__":
    main()
